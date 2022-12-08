# git stuff
# Import Module
import numpy as np
from tkinter import *
import RPi.GPIO as GPIO
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

GPIO.setmode(GPIO.BCM)
GPIO.setwarnings(False)
# GPIO.cleanup()

wb = load_workbook('DAK.xlsx')
ws = wb['Sheet2']

## Hardware ##
data = 2 #  26 DIOA on HV507, 14 SER on 74HC595 - Red
NBL = 3 #   29 !BL on HV507,  10 !SRCLR on 74HC595 - Green
NPol = 4 #  30 !POL on HV507, not present on 74HC595 - Clear
CLK  = 17 # 37 CLK on HV507,  11 SRCLK on 74HC595 - Blue
NLE = 27 #  38 !LE on HV507,  12 RCLK on 74HC595 - Yellow

HB1L = 19 # opto-isolator U4
HB1H = 26 # opto-isolator U5
HB2L = 6 #  opto-isolator U7
HB2H = 13 # opto-isolator U6
HB3L = 9 #  opto-isolator U8
HB3H = 11 # opto-isolator U9
HV100_ON = 20 # U3
HV300_ON = 21 # U2
DISCHARGE100V = 23 # MOSFET U10
DISCHARGE300V = 24 # MOSFET U11

GPIO.setup(data, GPIO.OUT, initial = 0)
GPIO.setup(NBL, GPIO.OUT, initial = 0)
GPIO.setup(NPol, GPIO.OUT, initial = 1)
GPIO.setup(CLK, GPIO.OUT, initial = 0)
GPIO.setup(NLE, GPIO.OUT, initial = 0)

GPIO.setup(HB1L, GPIO.OUT, initial = 0)
GPIO.setup(HB1H, GPIO.OUT, initial = 0)
GPIO.setup(HB2L, GPIO.OUT, initial = 0)
GPIO.setup(HB2H, GPIO.OUT, initial = 0)
GPIO.setup(HB3L, GPIO.OUT, initial = 0)
GPIO.setup(HB3H, GPIO.OUT, initial = 0)
GPIO.setup(HV100_ON, GPIO.OUT, initial = 0)
GPIO.setup(HV300_ON, GPIO.OUT, initial = 0)
GPIO.setup(DISCHARGE100V, GPIO.OUT, initial = 0)
GPIO.setup(DISCHARGE300V, GPIO.OUT, initial = 0)

size = 256
polarity = 1
RC_time_constant = 1 # delay for capcitor RC time constant
sleep = 0.01
#bit = 1

 
# Create Object
win = Tk()
 
# Add Title
win.title('NVMG')
 
# Add Geometry
# win.geometry("500x300")
 
# Keep track of the button state on/off
global is_on
is_on = False
 
# Define our switch function
running = False  # Global flag

def scanning():
    if running:  # Only do this if the Stop button has not been clicked
        polarity_swap()
    # After 1 second, call scanning again (create a recursive loop)
    win.after(1000, scanning)

def start():
    """Enable scanning by setting the global flag to True."""
    global running
    running = True
    
def stop():
    """Stop scanning by setting the global flag to False."""
    global running
    running = False
    

dak = np.array([
    [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [0,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
    [0,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],
    [0,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],
    [0,0,1,1,1,1,1,1,1,1,0,0,1,1,0,0],
    [0,0,0,1,1,1,1,1,1,0,0,0,0,1,1,0],
    [0,0,0,1,1,1,1,1,1,0,0,0,0,1,1,0],
    [0,0,1,1,1,1,1,1,1,1,0,0,1,1,0,0],
    [0,1,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [0,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],
    [0,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],])
    
letter_A = np.array([
    [0,0,0,0,0,1,1,1,1,1,1,0,0,0,0,0],
    [0,0,0,0,1,1,0,0,0,0,1,1,0,0,0,0],
    [0,0,0,1,1,0,0,0,0,0,0,1,1,0,0,0],
    [0,0,1,1,0,0,0,0,0,0,0,0,1,1,0,0],
    [0,1,1,0,0,0,0,0,0,0,0,0,0,1,1,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],])
    
letter_B = np.array([
    [1,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,1,1,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,1,1,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,1,1,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,1,1,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],])

letter_C = np.array([
    [0,0,0,1,1,1,1,1,1,1,1,1,1,0,0,0],
    [0,0,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [0,1,1,1,0,0,0,0,0,0,0,0,1,1,1,0],
    [1,1,1,0,0,0,0,0,0,0,0,0,0,1,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,1,0,0,0,0,0,0,0,0,0,0,1,1,1],
    [0,1,1,1,0,0,0,0,0,0,0,0,1,1,1,0],
    [0,0,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [0,0,0,1,1,1,1,1,1,1,1,1,1,0,0,0],])


letter_D = np.array([
    [1,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,1,1,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,1,1,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],])

letter_E = np.array([
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],])

letter_F = np.array([
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],
    [1,1,1,1,1,1,1,1,1,1,1,1,1,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],])

letter_G = np.array([
[0,0,0,1,1,1,1,1,1,1,1,1,1,0,0,0],
[0,0,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
[0,1,1,1,0,0,0,0,0,0,0,0,1,1,1,0],
[1,1,1,0,0,0,0,0,0,0,0,0,0,1,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
[1,1,0,0,0,0,0,0,0,0,1,1,1,1,1,1],
[1,1,0,0,0,0,0,0,0,0,1,1,1,1,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,1,0,0,0,0,0,0,0,0,0,0,1,1,1],
[0,1,1,1,0,0,0,0,0,0,0,0,1,1,1,0],
[0,0,1,1,1,1,1,1,1,1,1,1,1,1,0,0],
[0,0,0,1,1,1,1,1,1,1,1,1,1,0,0,0],])

letter_H = np.array([
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
[1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],
[1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1],])

display = letter_A
    
def inject_pattern():
    for row in range(12,16):
        for col in range(0,16):
            val = int(display[row,col])
            HV507_load_shift_register(val)
#             print("row = ", row," col = ", col, " val = ", val)
    for row in range(8,12):
        for col in range(0,16):
            val = int(display[row,col])
            HV507_load_shift_register(val)
#             print("row = ", row," col = ", col, " val = ", val)
    for row in range(4,8):
        for col in range(0,16):
            val = int(display[row,col])
            HV507_load_shift_register(val)
#             print("row = ", row," col = ", col, " val = ", val)
    for row in range(0,4):
        for col in range(0,16):                       
            val = int(display[row,col])
            HV507_load_shift_register(val)
#             print("row = ", row," col = ", col, " val = ", val)
    HV507_store_data_in_latches()
    return

def scroll_right():    
    global display
    display = np.roll(display,1, axis=1)
#     HV507_store_data_in_latches()
#     return display
    
def start_scrolling():
    global scrolling
    scrolling = True

def stop_scrolling():
    global scrolling
    scrolling = False

def scrolling():
    if scrolling:
        print('scroll right')
        scroll_right()
    win.after(1000, scrolling)


def close():
    print("entering close() function")
#     GPIO.cleanup()
    win.destroy()
    print("exiting close() function")
    return
    
def status():
    print("##### Status #####")
    print("data = ", GPIO.input(data))
    print("NBL = ", GPIO.input(NBL))
    print("CLK = ", GPIO.input(CLK))
    print("NLE = ", GPIO.input(NLE))
    print("NPol = ", GPIO.input(NPol))
    print("HB1L = ", GPIO.input(HB1L), " HB1H = ", GPIO.input(HB1H))
    print("HB2L = ", GPIO.input(HB2L), " HB2H = ", GPIO.input(HB2H))
    print("HB3L = ", GPIO.input(HB3L), " HB3H = ", GPIO.input(HB3H))
    print("HV100_ON = ", GPIO.input(HV100_ON))
    print("HV300_ON = ", GPIO.input(HV300_ON ))
    print("DISCHARGE100V = ", GPIO.input(DISCHARGE100V))
    print("DISCHARGE300V = ", GPIO.input(DISCHARGE300V))
    print("##################")
    return

def blank_display():
    for x in range(size):
        HV507_load_shift_register_low()
    HV507_store_data_in_latches()
    return
    
######################################################################################
########################### HV507 Function Table Functions ###########################
######################################################################################
    
def HV507_all_on():    
    GPIO.output(NBL, 0)
    GPIO.output(NPol, 0)
    return

def HV507_all_off():    
    GPIO.output(NBL, 0)
    GPIO.output(NPol, 1)
    return

def HV507_invert_mode():
    GPIO.output(NLE, 0)
    GPIO.output(NBL, 1)
    GPIO.output(NPol, 0)
    return

def HV507_load_shift_register(bit):
    GPIO.output(data, bit)
    GPIO.output(NLE, 0)
    GPIO.output(NBL, 1)
    GPIO.output(NPol, 1)
    GPIO.output(CLK, 1)
    GPIO.output(CLK, 0)
    return

def HV507_load_shift_register_low():
    GPIO.output(data, 0)
    GPIO.output(NLE, 0)
    GPIO.output(NBL, 1)
    GPIO.output(NPol, 1)
    GPIO.output(CLK, 1)
    GPIO.output(CLK, 0)
    return

def HV507_load_shift_register_high():
    GPIO.output(data, 1)
    GPIO.output(NLE, 0)
    GPIO.output(NBL, 1)
    GPIO.output(NPol, 1)
    GPIO.output(CLK, 1)
    GPIO.output(CLK, 0)
    return

def HV507_store_data_in_latches():    
    GPIO.output(NBL, 1)
    GPIO.output(NPol, 1)
    GPIO.output(NLE, 0)
    GPIO.output(NLE, 1) 
    return

def HV507_store_data_in_latches_not():    
    GPIO.output(NBL, 1)
    GPIO.output(NPol, 0)
    GPIO.output(NLE, 0)
    GPIO.output(NLE, 1) 
    return

######################################################################################
################################# H-Bridge Functions #################################
######################################################################################

def polarity_swap():
#     print("entering polarity_swap() function")
#     time.sleep(sleep)
    global polarity
    polarity = (1 - polarity)
    GPIO.output(NPol, polarity)
    if polarity == 1:
        HV_400v()
#         print("exiting polarity_swap() function with 300V and polarity of ", polarity)
        return
    HV_100v()
#     print("exiting polarity_swap() function with 100V and polarity of ", polarity)
    return

def HV_400v():
    print("entering HV_400v function")
    GPIO.output(HB1L, 0)
    GPIO.output(HB2H, 0)
    GPIO.output(HB3L, 0)
    GPIO.output(HB1H, 1)
    GPIO.output(HB2L, 1)    
    GPIO.output(HB3H, 1)
    print("exiting HV_400v function")
    return

def HV_100v():
    print("entering HV_100v function")
    GPIO.output(HB1H, 0)
    GPIO.output(HB2L, 0)
    GPIO.output(HB3H, 0)
    GPIO.output(HB1L, 1)
    GPIO.output(HB2H, 1)
    GPIO.output(HB3L, 1)    
    print("exiting HV_100v function")
    return

def picos_on():
    print("entering picos_on function")
    GPIO.output(HV100_ON, 1)
    GPIO.output(HV300_ON, 1)
    print("exiting picos_on function")
    return
    
def picos_off():
    print("entering picos_off function")
    GPIO.output(HV100_ON, 0)
    GPIO.output(HV300_ON, 0)
    print("exiting picos_off function")
    return

def power_up():
    print("entering power_up function")
    H_Bridge_float()
    picos_on()    
    HV_400v()
    print("exiting power_up function")
    return

def power_down():
    print("entering power_down function")    
    discharge()
#     H_Bridge_float()
    print("exiting power_down function")
    return

def H_Bridge_float():
    print("entering H_Bridge_float function")
    time.sleep(sleep)
    GPIO.output(HB1H, 0)
    time.sleep(sleep)
    GPIO.output(HB2H, 0)
    time.sleep(sleep)
    GPIO.output(HB3H, 0)
    time.sleep(sleep)
    GPIO.output(HB1L, 0)
    time.sleep(sleep)
    GPIO.output(HB2L, 0)
    time.sleep(sleep)
    GPIO.output(HB3L, 0)
    time.sleep(sleep)
    print("exiting H_Bridge_float function")
    return

def H_Bridge_ground():
    print("entering H_Bridge_ground function")
    time.sleep(sleep)
    GPIO.output(HB1H, 0)
    time.sleep(sleep)
    GPIO.output(HB2H, 0)
    time.sleep(sleep)
    GPIO.output(HB3H, 0)
    time.sleep(sleep)
    GPIO.output(HB1L, 1)
    time.sleep(sleep)
    GPIO.output(HB2L, 1)
    time.sleep(sleep)
    GPIO.output(HB3L, 1)
    time.sleep(sleep)
    print("exiting H_Bridge_ground function")
    return

def discharge():
    print("entering discharge function")
    picos_off()
    GPIO.output(DISCHARGE100V, 1)
    GPIO.output(DISCHARGE300V, 1)
    time.sleep(RC_time_constant)
    GPIO.output(DISCHARGE100V, 0)
    GPIO.output(DISCHARGE300V, 0)
    print("exiting discharge function")
    return

######################################################################################
###################################### Widgets ######################################
######################################################################################
powerUpButton = Button(win, text = "Power Up", command = power_up)
powerUpButton.pack()

powerDownButton = Button(win, text = "Power Down", command = power_down)
powerDownButton.pack()

blankDisplayButton = Button(win, text = "Blank Display", command = blank_display)
blankDisplayButton.pack()

injectpatternButton = Button(win, text = "Inject Pattern", command = inject_pattern)
injectpatternButton.pack()

# allOnButton = Button(win, text = "All On", command = HV507_all_on)
# allOnButton.pack()
# 
# allOffButton = Button(win, text = "All Off", command = HV507_all_off)
# allOffButton.pack()

# invertModeButton = Button(win, text = "Invert Mode", command = HV507_invert_mode)
# invertModeButton.pack()

loadShiftRegisterLowButton = Button(win, text = "Load Shift Register Low", command = HV507_load_shift_register_low)
loadShiftRegisterLowButton.pack()

loadShiftRegisterHighButton = Button(win, text = "Load Shift Register High", command = HV507_load_shift_register_high)
loadShiftRegisterHighButton.pack()

storeDataInLatchesButton = Button(win, text = "Store Data In Latches", command = HV507_store_data_in_latches)
storeDataInLatchesButton.pack()

storeDataInLatchesNotButton = Button(win, text = "Store Data In Latches NOT", command = HV507_store_data_in_latches_not)
storeDataInLatchesNotButton.pack()

statusButton = Button(win, text = "Status", command = status)
statusButton.pack()

polaritySwapButton = Button(win, text = "Polarity_Swap", command = polarity_swap)
polaritySwapButton.pack()

# oscilateButton = Button(win, text = "Oscilate", command = oscilate)
# oscilateButton.pack()
 
startButton = Button(win, text="Start Polarity Swap", command=start)
startButton.pack()

stopButton = Button(win, text="Stop Polarity Swap", command=stop)
stopButton.pack()

dischargeButton = Button(win, text="Discharge", command=discharge)
dischargeButton.pack()

exitButton = Button(win, text = "Exit Program", command = close)
exitButton.pack()

win.after(1000, scanning)  # After 1 second, call scanning

win.protocol("WM_DELETE_WINDOW", close) # exit cleanly

win.mainloop() # loop forever

